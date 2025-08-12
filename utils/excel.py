
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tempfile import NamedTemporaryFile

def generar_excel(nombre_establecimiento, fecha, df, responsable, supervisor, revision, nombre_reporte="Verificacion_BPM"):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_reporte.replace("_", " ")

    # Estilos
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")

    if nombre_reporte == "Verificacion_BPM":
        ws.append(["Nombre del establecimiento", nombre_establecimiento])
    ws.append(["Fecha", str(fecha)])
    ws.append([])

    # Encabezados de la tabla
    ws.append(list(df.columns))
    header_row = ws.max_row
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Filas de datos
    for row_idx, row in enumerate(df.values.tolist(), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Centrar selectores (C/NC) y observaciones a la izquierda
            if 3 <= col_idx <= 7:
                cell.alignment = center
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, max_length + 2)

    ws.append([])
    ws.append(["Responsable", responsable])
    if nombre_reporte == "Verificacion_BPM":
        ws.append(["Supervisor", supervisor])
        ws.append(["Revisión", revision])

    # Formato de nombre: nombre_reporte_YYYY-MM-DD.xlsx
    fecha_str = str(fecha)
    nombre_archivo = f"{nombre_reporte}_{fecha_str}.xlsx"
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=nombre_archivo.replace(" ", "_"))
    wb.save(tmp.name)
    return tmp.name
