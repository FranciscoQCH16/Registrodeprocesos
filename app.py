import streamlit as st
import pandas as pd
from utils.excel import generar_excel
from utils.dropbox import subir_a_dropbox
from datetime import date

# --- Autenticación usando secrets de Streamlit ---
USUARIOS_AUTORIZADOS = dict(st.secrets["auth"]) if "auth" in st.secrets else {}

def login():
    st.title("Iniciar sesión")
    usuario = st.text_input("Usuario")
    password = st.text_input("Contraseña", type="password")
    if st.button("Acceder"):
        if usuario in USUARIOS_AUTORIZADOS and password == USUARIOS_AUTORIZADOS[usuario]:
            st.session_state["autenticado"] = True
            try:
                st.rerun()
            except AttributeError:
                pass
        else:
            st.error("Usuario o contraseña incorrectos")

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if not st.session_state["autenticado"]:
    login()
    st.stop()

st.set_page_config(page_title="Reportes B.P.M.", layout="wide")

st.title("Sistema de Reportes L & D")

# Pestañas para distintos reportes
tabs_labels = [
    "Verificación de B.P.M.",
    "Formato Preparación de Soluciones D",
    "Registro de no conformidades y A. C.",
    "Registro y Control de Limpieza De Tanques",
    "Control de limpieza y desinfección Restaurante"
]
pestañas = st.tabs(tabs_labels)
with pestañas[4]:
    st.header("Control de limpieza y desinfección Restaurante UFPSO")

    st.subheader(":blue[Unir reportes diarios en un solo archivo semanal]")
    from datetime import timedelta
    import tempfile
    import re
    import openpyxl
    import io
    import dropbox


    # Áreas fijas por categoría (mover aquí para que esté disponible en todo el scope)
    categorias_areas = {
        "Baños": ["Paredes", "Pisos", "Lavamanos", "Inodoros"],
        "Almacen Fruver": ["Paredes", "Puertas", "Piso", "Estibas"],
        "Almacen Abarrotes": ["Paredes", "Pisos", "Puerta", "Estibas", "Estantes"],
        "Cavas de almacenamiento/nevera": ["Paredes interiores", "Pisos", "Puertas", "Exteriores"],
        "Cocina caliente": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Panadería Y Pastelería": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Línea de servicio": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Comedor": ["Iluminarias", "Paredes", "Pisos", "Puertas", "Mesones","Ventanas"],
        "Acopio de residuos": ["Orgánicos", "Inorgánicos", "Puertas", "Canecas","Tapas","Trampas de grasa"]
    }

    # Selección de rango de fechas (semana)
    col1, col2 = st.columns(2)
    fecha_inicio = col1.date_input("Fecha de inicio de la semana", value=date.today() - timedelta(days=date.today().weekday()), key="fecha_inicio_semana")
    fecha_fin = col2.date_input("Fecha de fin de la semana", value=date.today(), key="fecha_fin_semana")

    if st.button("Unir reportes diarios de la semana y subir a Dropbox", key="btn_unir_reportes_semanal"):
        from utils.dropbox import get_access_token, DROPBOX_FOLDER
        access_token = get_access_token()
        dbx = dropbox.Dropbox(oauth2_access_token=access_token)
        archivos = dbx.files_list_folder(DROPBOX_FOLDER).entries
        patron = re.compile(r"Control_Limpieza_Restaurante_UFPSO_(\d{4}-\d{2}-\d{2})\.xlsx", re.IGNORECASE)
        archivos_semana = []
        for entry in archivos:
            if hasattr(entry, 'name'):
                m = patron.search(entry.name)
                if m:
                    fecha_archivo = m.group(1)
                    if fecha_inicio <= date.fromisoformat(fecha_archivo) <= fecha_fin:
                        archivos_semana.append((entry.name, fecha_archivo))
        if not archivos_semana:
            st.warning("No se encontraron reportes diarios para ese rango de fechas.")
        else:
            # Agrupar por fecha real (YYYY-MM-DD)
            tablas_por_fecha = {}
            fechas_ordenadas = []
            for nombre_archivo, fecha_archivo in sorted(archivos_semana, key=lambda x: x[1]):
                ruta = DROPBOX_FOLDER + nombre_archivo
                _, res = dbx.files_download(ruta)
                with io.BytesIO(res.content) as f:
                    wb = openpyxl.load_workbook(f)
                    ws = wb.active
                    data = list(ws.values)
                    data = [row for row in data if any(row)]
                    bloques = []
                    i = 0
                    while i < len(data):
                        # Buscar fila de título de categoría
                        if data[i] and all(isinstance(x, str) and x.strip() != '' for x in data[i][0:1]) and data[i][0] in [*categorias_areas.keys()]:
                            categoria = data[i][0]
                            # Encabezados
                            if i+1 < len(data) and "Área" in data[i+1]:
                                headers = data[i+1]
                                # Filas de datos
                                j = i+2
                                filas = []
                                while j < len(data) and (data[j][0] not in categorias_areas.keys() and any(data[j])):
                                    filas.append(data[j])
                                    j += 1
                                # Guardar bloque: primero la fila de título, luego encabezados, luego datos
                                bloques.append([list(data[i])] + [list(headers)] + [list(f) for f in filas])
                                i = j
                            else:
                                i += 1
                        else:
                            i += 1
                    # Unir todos los bloques en una sola lista de filas
                    filas_final = []
                    for bloque in bloques:
                        filas_final.extend(bloque)
                        filas_final.append(["" for _ in range(3)])  # Separador
                    # Convertir a DataFrame (rellenar filas cortas)
                    max_cols = max(len(f) for f in filas_final) if filas_final else 3
                    filas_final = [f + [""]*(max_cols-len(f)) for f in filas_final]
                    df = pd.DataFrame(filas_final)
                    if fecha_archivo not in tablas_por_fecha:
                        tablas_por_fecha[fecha_archivo] = []
                        fechas_ordenadas.append(fecha_archivo)
                    tablas_por_fecha[fecha_archivo].append(df)
            # Unir por fecha y guardar en Excel lado a lado
            from openpyxl.styles import Font
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="Control_Limpieza_Restaurante_UFPSO_Semanal_") as tmp:
                with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
                    col_offset = 0
                    for fecha in fechas_ordenadas:
                        df_fecha = pd.concat(tablas_por_fecha[fecha], ignore_index=True)
                        df_fecha.to_excel(writer, sheet_name="Semana", startrow=0, startcol=col_offset, index=False)
                        ws = writer.sheets["Semana"]
                        ws.cell(row=1, column=col_offset+1, value=fecha)
                        # Ajustar ancho de columnas automáticamente
                        for i, col in enumerate(df_fecha.columns, start=col_offset+1):
                            max_length = max([len(str(cell)) if cell is not None else 0 for cell in df_fecha[col]])
                            header = str(col)
                            max_length = max(max_length, len(header))
                            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_length + 2
                        # Poner en negrita los títulos de categoría y encabezados
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_offset+1, max_col=col_offset+len(df_fecha.columns)):
                            if row[0].value in categorias_areas.keys() or row[0].value == "Área":
                                for cell in row:
                                    cell.font = Font(bold=True)
                        col_offset += len(df_fecha.columns) + 2
                nombre_archivo_final = f"Control_Limpieza_Restaurante_UFPSO_Semanal_{fecha_inicio}_a_{fecha_fin}.xlsx"
                ruta_destino = DROPBOX_FOLDER + nombre_archivo_final
                with open(tmp.name, 'rb') as f:
                    dbx.files_upload(f.read(), ruta_destino, mode=dropbox.files.WriteMode.overwrite)
                try:
                    shared_link_metadata = dbx.sharing_create_shared_link_with_settings(ruta_destino)
                    url = shared_link_metadata.url
                except dropbox.exceptions.ApiError as e:
                    if (hasattr(e, 'error') and hasattr(e.error, 'is_shared_link_already_exists') and e.error.is_shared_link_already_exists()):
                        links = dbx.sharing_list_shared_links(path=ruta_destino, direct_only=True).links
                        url = links[0].url if links else None
                    else:
                        raise
                if url:
                    st.success(f"Archivo semanal generado y subido. Acceso: {url}")
            # Mostrar tablas lado a lado en Streamlit
            cols = st.columns(len(fechas_ordenadas))
            for idx, fecha in enumerate(fechas_ordenadas):
                df_fecha = pd.concat(tablas_por_fecha[fecha], ignore_index=True)
                cols[idx].markdown(f"**{fecha}**")
                cols[idx].dataframe(df_fecha)

    fecha_control = st.date_input("Fecha del control", value=date.today(), key="fecha_control_ufpso", help="Fecha")

    # Áreas fijas por categoría
    categorias_areas = {
        "Baños": ["Paredes", "Pisos", "Lavamanos", "Inodoros"],
        "Almacen Fruver": ["Paredes", "Puertas", "Piso", "Estibas"],
        "Almacen Abarrotes": ["Paredes", "Pisos", "Puerta", "Estibas", "Estantes"],
        "Cavas de almacenamiento/nevera": ["Paredes interiores", "Pisos", "Puertas", "Exteriores"],
        "Cocina caliente": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Panadería Y Pastelería": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Línea de servicio": ["Maquinaria y equipos", "Paredes", "Pisos", "Puertas", "Mesones"],
        "Comedor": ["Iluminarias", "Paredes", "Pisos", "Puertas", "Mesones","Ventanas"],
        "Acopio de residuos": ["Orgánicos", "Inorgánicos", "Puertas", "Canecas","Tapas","Trampas de grasa"]
    }
    # Selector único de día para todo el reporte (desplegable)
    dias_opciones = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    dia_seleccionado = st.selectbox("Selecciona el día del reporte", dias_opciones, index=0)
    columnas_base = ["Área", dia_seleccionado, "Observaciones"]
    dfs_categorias = {}
    for nombre, key in [(k, k.lower().replace(' ', '_')) for k in categorias_areas.keys()]:
        st.subheader(nombre)
        areas = categorias_areas[nombre]
        data = []
        for i, area in enumerate(areas):
            cols = st.columns([2, 1, 2])
            # Área como texto fijo
            cols[0].markdown(f"<div style='padding-top:0.5em'>{area}</div>", unsafe_allow_html=True)
            # Selector R/NR para el día seleccionado (desplegable)
            val = cols[1].selectbox(
                dia_seleccionado,
                ["Seleccione si se realizó", "R", "NR"],
                key=f"{key}_{i}_dia",
                label_visibility="collapsed",
                index=0,
                help=f"Seleccione R o NR para {dia_seleccionado}"
            )
            val = val if val in ["R", "NR"] else ""
            observ = cols[2].text_input("Observaciones", key=f"{key}_obs_{i}", label_visibility="collapsed", placeholder="Observaciones")
            data.append([area, val, observ])
        df = pd.DataFrame(data, columns=columnas_base)
        dfs_categorias[nombre] = df

    responsable_general = st.text_input("Responsable general del reporte", key="responsable_general_restaurante")
    if st.button("Generar y subir reporte de limpieza Restaurante a Dropbox", key="btn_reporte_restaurante"):
        # Construir una super tabla con títulos de categoría como filas separadoras
        super_tabla = []
        for nombre in categorias_areas.keys():
            # Fila de título de categoría
            super_tabla.append([f"{nombre}"] + ["" for _ in range(2)])
            # Encabezados de la tabla
            super_tabla.append(list(dfs_categorias[nombre].columns))
            # Filas de la tabla
            for fila in dfs_categorias[nombre].values:
                super_tabla.append(list(fila))
            # Fila vacía para separar categorías
            super_tabla.append(["" for _ in range(3)])
        df_super = pd.DataFrame(super_tabla)
        archivo = generar_excel(
            "CONTROL DE LIMPIEZA Y DESINFECCION RESTAURANTE UFPSO",
            fecha_control,
            df_super,
            responsable_general,
            "",
            "",
            nombre_reporte="Control_Limpieza_Restaurante_UFPSO"
        )
        url = subir_a_dropbox(archivo)
        st.success(f"Reporte generado y subido. Acceso: {url}")

    # (render_categoria eliminado, ya no se llama aquí)
with pestañas[3]:
    st.header("Registro y control de limpieza y desinfección de tanques de almacenamiento de agua")
    columnas4 = [
        "Fecha",
        "Actividad",
        "Programado por",
        "Responsable",
        "Cumple o no cumple"
    ]
    filas4 = st.number_input("Cantidad de registros", min_value=1, max_value=20, value=3, key="filas4")
    cols_titulos4 = st.columns(len(columnas4))
    for j, col in enumerate(columnas4):
        cols_titulos4[j].markdown(f"**{col}**")
    data4 = []
    for i in range(filas4):
        cols4 = st.columns(len(columnas4))
        fila4 = []
        for j, col in enumerate(columnas4):
            if j == 0:  # Fecha
                val = cols4[j].date_input(col, key=f"date4_{i}_{j}", label_visibility="collapsed")
            elif j == 4:  # Cumple o no cumple
                val = cols4[j].selectbox(col, ["Seleccione", "Cumple", "No cumple"], key=f"sel4_{i}_{j}", label_visibility="collapsed", index=0)
                val = val if val in ["Cumple", "No cumple"] else ""
            else:
                val = cols4[j].text_input(col, key=f"txt4_{i}_{j}", label_visibility="collapsed", placeholder=col)
            fila4.append(val)
        data4.append(fila4)
    df4 = pd.DataFrame(data4, columns=columnas4)

    responsable4 = st.text_input("Responsable del registro", key="responsable4")

    if st.button("Generar y subir reporte de limpieza de tanques a Dropbox"):
        archivo4 = generar_excel(
            "",
            date.today(),
            df4,
            responsable4,
            "",
            "",
            nombre_reporte="Registro_Limpieza_Tanques"
        )
        url4 = subir_a_dropbox(archivo4)
        st.success(f"Reporte generado y subido. Acceso: {url4}")
with pestañas[2]:
    st.header("Registro de no conformidades y acciones correctivas")
    columnas3 = [
        "Fecha de la incidencia",
        "Area/Equipo afectado",
        "Descripcion de la No Conformidad",
        "Accion correctiva inmediata aplicada",
        "Responsable",
        "Verificacion",
        "Fecha de verificacion"
    ]
    filas3 = st.number_input("Cantidad de incidencias", min_value=1, max_value=20, value=3, key="filas3")
    cols_titulos3 = st.columns(len(columnas3))
    for j, col in enumerate(columnas3):
        cols_titulos3[j].markdown(f"**{col}**")
    data3 = []
    for i in range(filas3):
        cols3 = st.columns(len(columnas3))
        fila3 = []
        for j, col in enumerate(columnas3):
            if j == 0 or j == 6:  # Fechas
                val = cols3[j].date_input(col, key=f"date3_{i}_{j}", label_visibility="collapsed")
            else:
                val = cols3[j].text_input(col, key=f"txt3_{i}_{j}", label_visibility="collapsed", placeholder=col)
            fila3.append(val)
        data3.append(fila3)
    df3 = pd.DataFrame(data3, columns=columnas3)

    responsable3 = st.text_input("Responsable del registro", key="responsable3")

    if st.button("Generar y subir reporte de no conformidades a Dropbox"):
        archivo3 = generar_excel(
            "",
            date.today(),
            df3,
            responsable3,
            "",
            "",
            nombre_reporte="Registro_No_Conformidades"
        )
        url3 = subir_a_dropbox(archivo3)
        st.success(f"Reporte generado y subido. Acceso: {url3}")

with pestañas[0]:
    st.header("Verificación de B.P.M.")
    nombre_establecimiento = st.text_input("Nombre del establecimiento")
    fecha = st.date_input("Fecha", value=date.today())

    st.subheader("Datos de Operarios")
    columnas = [
        "Nombre del operario", "Lavado de manos", "Dotación", "Utilizan toda la dotación",
        "Uñas cortas limpias y sin esmalte", "Accesorios y presentación personal", "Observaciones"
    ]
    # Para los selectores, la primera opción es un hint con el nombre de la columna
    opciones_selector = lambda col: [f"Seleccione {col}", "C", "NC"]
    data = []
    operarios = ["Juan Carlos", "Humberto Florez", "Francisco Quintero"]
    filas = len(operarios)
    # Mostrar títulos de columna
    cols_titulos = st.columns(len(columnas))
    for j, col in enumerate(columnas):
        cols_titulos[j].markdown(f"**{col}**")
    # Inputs de la tabla
    for i, nombre_operario in enumerate(operarios):
        cols = st.columns(len(columnas))
        fila = []
        for j, col in enumerate(columnas):
            if j == 0:
                # Nombre del operario fijo
                cols[j].markdown(f"<div style='padding-top:0.5em'>{nombre_operario}</div>", unsafe_allow_html=True)
                fila.append(nombre_operario)
            elif 1 <= j <= 5:
                val = cols[j].selectbox(
                    col,
                    opciones_selector(col),
                    key=f"sel_{i}_{j}",
                    label_visibility="collapsed",
                    index=0,
                    help=f"Seleccione C o NC para {col}"
                )
                val = val if val in ["C", "NC"] else ""
                fila.append(val)
            else:
                val = cols[j].text_input(
                    col, key=f"txt_{i}_{j}", label_visibility="collapsed", placeholder=col
                )
                fila.append(val)
        data.append(fila)
    df = pd.DataFrame(data, columns=columnas)

    st.subheader("Firmas")
    responsable = st.text_input("Responsable")
    supervisor = st.text_input("Supervisor")
    revision = st.text_input("Revisión")

    if st.button("Generar y subir reporte a Dropbox"):
        archivo = generar_excel(
            nombre_establecimiento,
            fecha,
            df,
            responsable,
            supervisor,
            revision,
            nombre_reporte="Verificacion_BPM"
        )
        url = subir_a_dropbox(archivo)
        st.success(f"Reporte generado y subido. Acceso: {url}")

with pestañas[1]:
    st.header("Formato de Registro de Preparación de Soluciones Desinfectantes")
    fecha_registro = st.date_input("Fecha del registro", value=date.today(), key="fecha_registro")

    columnas2 = [
        "Fecha", "Desinfectante", "Concentracion preparada", "Responsable", "Accion correctiva"
    ]
    filas2 = st.number_input("Cantidad de registros", min_value=1, max_value=20, value=3, key="filas2")
    cols_titulos2 = st.columns(len(columnas2))
    for j, col in enumerate(columnas2):
        cols_titulos2[j].markdown(f"**{col}**")
    data2 = []
    for i in range(filas2):
        cols2 = st.columns(len(columnas2))
        fila2 = []
        for j, col in enumerate(columnas2):
            if j == 0:  # Columna Fecha
                val = cols2[j].date_input(col, key=f"date2_{i}_{j}", label_visibility="collapsed")
            else:
                val = cols2[j].text_input(col, key=f"txt2_{i}_{j}", label_visibility="collapsed", placeholder=col)
            fila2.append(val)
        data2.append(fila2)
    df2 = pd.DataFrame(data2, columns=columnas2)

    responsable2 = st.text_input("Responsable del registro", key="responsable2")

    if st.button("Generar y subir reporte de soluciones a Dropbox"):
        archivo2 = generar_excel(
            "",
            fecha_registro,
            df2,
            responsable2,
            "",
            "",
            nombre_reporte="Registro_Preparacion_Soluciones_Desinfectantes"
        )
        url2 = subir_a_dropbox(archivo2)
        st.success(f"Reporte generado y subido. Acceso: {url2}")